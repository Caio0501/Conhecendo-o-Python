import openpyxl
import random

wb = openpyxl.Workbook()

ws = wb.active

ws.cell(1, 1, "Alunos")
ws.cell(1, 2, "BIM 1")
ws.cell(1, 3, "BIM 2")
ws.cell(1, 4, "BIM 3")
ws.cell(1, 5, "BIM 5")

for num in range(10):
    ws.cell(num + 2, 1, "Aluno" + str(num + 1))

for num in range(10):
    for bim in range(4):
        ws.cell(num + 2, bim + 2, random.randint(1, 10))

wb.save("Excel.xlsx")
wb.close()
